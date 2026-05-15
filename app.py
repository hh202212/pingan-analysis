import streamlit as st
import pdfplumber
import pandas as pd
import io
import os
import re
from openpyxl import load_workbook

# 1. 页面设置
st.set_page_config(page_title="钱坤大挪移-资产迁移系统", layout="wide")

# CSS 注入：平安橙色调 + 强制思源黑体
st.markdown("""
    <style>
    :root { --pa-orange: #FF6600; }
    .main { background-color: #fcfcfc; }
    .stButton>button { 
        background-color: var(--pa-orange); 
        color: white; 
        border-radius: 8px; 
        width: 100%;
        height: 3em;
        font-weight: bold;
    }
    .stDownloadButton>button { 
        background-color: #2E7D32; 
        color: white; 
        border-radius: 8px;
    }
    html, body, [class*="css"], h1, h2, h3, div, span, p, td, th { 
        font-family: 'Source Han Sans SC', '思源黑体', sans-serif !important; 
    }
    h1, h2, h3 { color: var(--pa-orange); }
    .disclaimer-box { 
        font-size: 13px; 
        color: #555; 
        margin-top: 30px; 
        border: 1px solid #ffccbc; 
        background-color: #fff5f2;
        padding: 15px; 
        border-radius: 10px;
        line-height: 1.6;
    }
    .preview-header {
        background-color: var(--pa-orange);
        color: white;
        padding: 10px;
        border-radius: 5px 5px 0 0;
        text-align: center;
        font-weight: bold;
    }
    </style>
    """, unsafe_allow_html=True)

st.title("🛡️ 钱坤大挪移-资产迁移系统 (V8.5 稳定版)")

# --- 核心辅助：分离方案并只取纯数据 (修复云端 NoneType 报错) ---
def extract_all_schemes(uploaded_file):
    all_schemes = []
    current_scheme_data = []
    
    # 【核心修复】：将 Streamlit 的上传对象强制转化为纯内存字节流，彻底杜绝指针错误
    pdf_bytes = io.BytesIO(uploaded_file.getvalue())
    
    try:
        with pdfplumber.open(pdf_bytes) as pdf:
            # 防崩拦截：如果 PDF 解析失败返回空，直接退出
            if pdf is None or not hasattr(pdf, 'pages'):
                return []
                
            for page in pdf.pages:
                tables = page.find_tables()
                if not tables: continue
                for t_obj in tables:
                    data = t_obj.extract()
                    if not data: continue
                    
                    # 寻找纯数据行起始（第一列是数字）
                    data_start = -1
                    first_year = -1
                    for r_idx, row in enumerate(data):
                        if row[0] and str(row[0]).strip().isdigit():
                            data_start = r_idx
                            first_year = int(str(row[0]).strip())
                            break
                    
                    if data_start != -1:
                        is_new_scheme = (first_year == 1)
                        
                        if is_new_scheme:
                            # 结算上一个方案
                            if current_scheme_data:
                                all_schemes.append(current_scheme_data)
                            # 开启新方案，只拿数据行
                            current_scheme_data = data[data_start:]
                        elif current_scheme_data is not None:
                            # 续表，直接拼接数据行
                            current_scheme_data.extend(data[data_start:])
                            
            if current_scheme_data:
                all_schemes.append(current_scheme_data)
                
    except Exception as e:
        st.error(f"PDF 底层解析失败: {e}")
        
    return all_schemes

# --- 2. 界面布局 ---
with st.sidebar:
    st.header("📋 业务参数录入")
    cust_name = st.text_input("客户姓名", value="客户")
    cust_age = st.number_input("客户年龄", value=45)
    principal = st.number_input("拟迁移总资产 (元)", value=400000)
    bank_rate = st.number_input("假定定存利率 (%)", value=0.95) / 100
    
    st.write("---")
    pdf_file = st.file_uploader("📤 第一步：上传 PDF 建议书", type="pdf")
    
    template_path = "template.xlsx"
    has_template = os.path.exists(template_path)
    
    if has_template:
        st.success("✅ 核心模板已加载 (template.xlsx)")
    else:
        st.error("❌ 仓库中未找到 template.xlsx 模板")

    start_calc = st.button("🚀 第二步：执行钱坤大挪移")

# --- 3. 后台计算逻辑 ---
if start_calc and pdf_file and has_template:
    try:
        with st.spinner('⏳ 正在解析建议书，精准锁定方案 2 数据...'):
            all_schemes = extract_all_schemes(pdf_file)
            
            if not all_schemes:
                st.error("未能识别到 PDF 里的纯数据行，请确认 PDF 文件是否损坏或加密。")
                st.stop()
            
            # 精准提取方案 2
            if len(all_schemes) >= 2:
                matrix_data = all_schemes[1]  # 索引 1 对应方案 2
                st.toast("🎯 已成功锁定并提取【方案 2】数据")
            else:
                matrix_data = all_schemes[0]  # 如果只有一个方案则降级使用方案 1
                st.toast("⚠️ PDF 中仅识别到一个方案，已默认提取【方案 1】")
            
            wb = load_workbook(template_path)
            
            # 填充“原表”参数
            if "原表" in wb.sheetnames:
                ws_raw = wb["原表"]
                ws_raw["D1"] = cust_age
                ws_raw["H1"] = principal
                ws_raw["D2"] = bank_rate
            
            # 填充“贴主险建议书”数据
            if "贴主险建议书" in wb.sheetnames:
                ws_paste = wb["贴主险建议书"]
                start_row = 5 
                for r_idx, row_data in enumerate(matrix_data):
                    for c_idx, val in enumerate(row_data):
                        # 清洗并排雷：防止 'None' 毒害 Excel 公式
                        if val is None or str(val).strip() == "" or str(val).lower() == "none":
                            clean_val = ""
                        else:
                            try:
                                clean_val = float(str(val).replace(',', '').replace(' ', ''))
                            except:
                                clean_val = str(val).replace('\n', ' ')
                        ws_paste.cell(row=start_row + r_idx, column=c_idx + 1, value=clean_val)
            
            # C. 抓取“建议展示”页数据供网页预览
            preview_data = []
            if "建议展示" in wb.sheetnames:
                ws_show = wb["建议展示"]
                for row in ws_show.iter_rows(min_row=1, max_row=35, values_only=True):
                    preview_data.append(row)
            
            # D. 导出文件
            output_excel = io.BytesIO()
            wb.save(output_excel)
            
            # --- 4. 预览与展示 ---
            st.balloons()
            st.success(f"🎉 {cust_name} 先生/女士，您的资产迁移方案（基于方案 2）已生成！")
            
            st.markdown('<div class="preview-header">📊 资产方案建议展示 (核心数据预览)</div>', unsafe_allow_html=True)
            
            if preview_data:
                df_show = pd.DataFrame(preview_data)
                df_show = df_show.fillna("")
                st.dataframe(df_show, use_container_width=True, height=600)
            
            st.markdown("""
                <div class="disclaimer-box">
                <strong>💡 钱坤大挪移测算模型温馨提示：</strong><br>
                1. <strong>数据有效性：</strong>本测算基于您上传的官方建议书及录入的银行假定利率。实际利益请以保险合同及每月官方结算利率为准。<br>
                2. <strong>万能账户：</strong>演示中高于保证利率的部分为非保证利益，实际收益随市场波动。<br>
                3. <strong>对比逻辑：</strong>银行存款对比仅作资产配置参考，协助理解收益及流动性差异，不构成投资建议。<br>
                4. <strong>版权说明：</strong>本工具生成内容仅供内部培训与辅助演示，严禁公开发布。
                </div>
            """, unsafe_allow_html=True)
            
            st.write("---")
            col1, col2 = st.columns(2)
            with col1:
                st.download_button(
                    label="📥 下载测算建议书 Excel (自动计算公式)",
                    data=output_excel.getvalue(),
                    file_name=f"{cust_name}_资产迁移建议书.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            with col2:
                st.info("🖼️ 提示：点击左侧下载 Excel，使用 Office/WPS 打开后公式将自动激活，您可以直接框选“建议展示”区域另存为高清图片供客户查阅。")

    except Exception as e:
        st.error(f"❌ 运行异常，请检查模板结构: {str(e)}")
